import private.DB as DB
import private.tokens
import sys
from openpyxl import Workbook
from openpyxl import load_workbook
from notion.client import NotionClient
from notion.block import FileBlock


conn = DB.get_connection_real()
cur = DB.get_cursor(conn)

day_off_origin_time = ['종일', '오전 반차', '오후 반차']


def day_off_raw():

    excel_document = load_workbook(filename='./day_off_raw.xlsx')
    sheet = excel_document['Sheet1']

    wb = Workbook()
    ws = wb.active

    row_count = len(sheet['D'])

    company_id = input('등록을 요청한 회사 아이디를 입력해주세요.: ')

    SQL = ''
    cur.execute(SQL)
    auth_id = cur.fetchall().pop()['UserID']

    for num in range(2, row_count+1):
        employee_name = str(sheet['A%s' %num].value) #직원 이름
        day_off_date = str(sheet['B%s' %num].value) #연차등록요청날짜
        day_off_time = str(sheet['C%s' %num].value) #종일/반차 등
        if day_off_time in day_off_origin_time:
            pass
        else:
            print('유효하지 않은 연차 종류가 포함되어있습니다. 시스템을 종료합니다.')
            sys.exit()
        day_off_type = str(sheet['D%s' %num].value) #연차종류
        split_rocation = day_off_type.find(' ')
        day_off_type = day_off_type[:split_rocation]
        if day_off_type == '연차':
            day_off_type = '1'
        else:
            day_off_type = '0'

        day_off_reason = str(sheet['E%s' %num].value) #연차 요청 사유
        if day_off_reason is 'None':
            day_off_reason = '[Jober] 초기 세팅'

    wb.save(filename='./%s.xlsx' % company_id)

    company_info = [company_id, company_name]

    return company_info


def notion_connect(company_info):
    client = NotionClient(token_v2=private.tokens.notion)

    url = ''
    cv = client.get_collection_view(url)

    row = cv.collection.add_row()
    row.name = company_info[1] + '의 연차 초기 셋팅'

    page = client.get_block(url_or_id=row.id)

    child = page.children.add_new(FileBlock)
    child.upload_file('./html/%s.xlsx' % company_info[0])

company_info = day_off_raw()
notion_connect(company_info)

